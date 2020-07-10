#!/usr/bin/env python
# coding: utf-8

# In[1]:


from selenium import webdriver
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time


# In[2]:


browser = webdriver.Firefox(executable_path=r'C:\Users\AJ\Documents\Pyhton-whatsapp-auto\geckodriver-v0.26.0-win64\geckodriver.exe')


# In[29]:


from openpyxl import Workbook, load_workbook


# In[30]:


wb = load_workbook('Live_Processes.xlsx')
print(wb.sheetnames)
work = wb['Model 1 & 2']


# In[33]:


c = work['AE209']
print(c.value)
start = 209
end = 416


# In[36]:


while start <= end:
    cellName = 'AE'+ str(start)
    c = work[cellName]
    url = c.value
    comment = ''
    try:
        browser.get(url)
        wait = WebDriverWait(browser, 500);#5 sec Wait to scan the QR manually

        spans = browser.find_elements_by_class_name("signalFieldValue_c079fcf3")
        #print(len(spans))


        sdd = 0
        pdd = 0
        uat = 0
        gfcf = 0
        kfas = 0

        kfaCell = 'AM' + str(start)
        e = work[kfaCell]
        if e.value.lower == 'yes':
            kfas = 1

        for ele in spans:
            fileName = ele.get_attribute('innerText')
            n = len(fileName) - 31
            finalName = fileName[0:n].lower()
            if finalName.find('pdd'):
                ext = finalName[finalName.find('.'):n]
                if ext == '.docx' or ext == '.doc':
                    pdd = 1
            if finalName.find('sdd'):
                sdd = 1
            if finalName.find('uat'):
                if finalName.find('Sign_Off'):
                    uat = 1
            if kfas == 1:
                if finalName.find('gfcf'):
                    gfcf = 1

        if pdd == 0:
            comment = comment+"PDD Not Found, "
        if sdd == 0:
            comment = comment+"SDD Not Found, "
        if uat == 0:
            comment = comment+"UAT Sign Off Not Found, "
        if kfas == 1 and gfcf == 0:
            comment = comment+"GFCF Approval Mail Not Found, "
    except:
        comment = "Bad Url"
    writeCell = 'BG' + str(start)
    c4 = work[writeCell]
    c4.value = comment
    wb.save("Live_Processes.xlsx")
    start += 1


# In[27]:


name = '''Business UAT Sign Off_e-Invoice Tesco.msg
Press C to open file hover card'''
print(len(name))
n = len(name) - 31
name = name[0:n].lower()
print(name)
print(name.find('Sign Off'))
print(name[name.find('.'):n])


# In[32]:





# In[ ]:




